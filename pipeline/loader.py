# -*- coding: utf-8 -*-
"""
数据加载器：机构、学校、问卷星
"""

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from .config import (
    INSTITUTION_JSON,
    QUESTIONNAIRE_XLSX,
    SCHOOL_DIR,
    SCHOOL_FILE_PATTERN,
)


def load_institutions(path: Optional[Path] = None) -> List[Dict[str, Any]]:
    """加载托育机构平台注册备案数据"""
    path = path or INSTITUTION_JSON
    if not path.exists():
        return []
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, list) else []


def parse_zoning(zoning_name: str) -> Tuple[str, str]:
    """
    从 zoning_name 解析省、市（或直辖市下的区县）
    例：云南省大理白族自治州大理市 -> (云南省, 大理白族自治州)
    例：陕西省西安市莲湖区 -> (陕西省, 西安市)
    例：北京市市辖区西城区 -> (北京市, 西城区)
    例：北京市辖区昌平区 -> (北京市, 昌平区)
    """
    if not zoning_name or not isinstance(zoning_name, str):
        return "", ""
    s = zoning_name.strip()
    province, city = "", ""
    # 省级：直辖市/省/自治区
    for sep in ["自治区", "省", "市"]:
        idx = s.find(sep)
        if idx >= 0:
            province = s[: idx + len(sep)]
            rest = s[idx + len(sep) :].lstrip()
            break
    if not province and s:
        province = s
        return province, ""
    if not rest:
        return province, ""

    # 直辖市：北京市/天津市/上海市/重庆市 等，中间常有 辖区 或 市辖区
    if province in ("北京市", "天津市", "上海市", "重庆市"):
        for prefix in ("市辖区", "辖区"):
            if rest.startswith(prefix):
                rest = rest[len(prefix) :].lstrip()
                break
        # 取区县：昌平区、西城区、朝阳区 等
        if rest:
            for suffix in ("区", "县"):
                if suffix in rest:
                    parts = rest.split(suffix)
                    if len(parts) >= 2 and parts[0]:
                        city = parts[0] + suffix
                        break
            if not city and rest:
                city = rest
        return province, city or rest

    # 普通省份：市级（自治州/地区/盟/市）或区县
    for sep in ["自治州", "地区", "盟", "市"]:
        idx = rest.find(sep)
        if idx >= 0:
            city = rest[: idx + len(sep)]
            return province, city
    if "区" in rest:
        city = rest.split("区")[0] + "区"
    elif "县" in rest:
        city = rest.split("县")[0] + "县"
    else:
        city = rest
    return province, city


def infer_host_type(item: Dict[str, Any]) -> str:
    """
    从 institution_type 推断举办主体
    事业单位 -> 公办；营利性、非营利性 -> 民办
    """
    t = (item.get("institution_type") or "").strip()
    if t == "事业单位":
        return "公办"
    if t in ("营利性", "非营利性"):
        return "民办"
    return "未知"


def infer_is_puhui(item: Dict[str, Any]) -> str:
    """
    从机构名称推断是否普惠（按名称关键词推断）
    """
    text = " ".join(
        filter(
            None,
            [
                str(item.get("institution_name") or ""),
                str(item.get("institution_other_name") or ""),
            ],
        )
    )
    return "普惠" if "普惠" in text else "非普惠"


def infer_service_modes(item: Dict[str, Any]) -> List[str]:
    """
    从机构名称推断服务模式（按名称关键词推断，可多选）
    全日托、半月托/半日托、小时托/计时托
    """
    text = " ".join(
        filter(
            None,
            [
                str(item.get("institution_name") or ""),
                str(item.get("institution_other_name") or ""),
            ],
        )
    )
    modes = []
    if any(k in text for k in ("全日托", "全日制")):
        modes.append("全日托")
    if any(k in text for k in ("半月托", "半日托")):
        modes.append("半月托")
    if any(k in text for k in ("小时托", "计时托")):
        modes.append("小时托")
    return modes


def load_schools(
    dir_path: Optional[Path] = None,
    pattern: Optional[str] = None,
    region_filter: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    加载托育学校注册备案数据
    记录格式：多行一组，用 ^_^ 分隔
    region_filter: 若指定，只加载 moe_majors_{region_filter}_*.txt，加速
    """
    dir_path = dir_path or SCHOOL_DIR
    if not dir_path.exists():
        return []

    if pattern:
        files = sorted(dir_path.glob(pattern))
    elif region_filter:
        files = sorted(dir_path.glob(f"moe_majors_{region_filter}_*.txt"))
    else:
        files = sorted(dir_path.glob(SCHOOL_FILE_PATTERN))

    records: List[Dict[str, Any]] = []

    for fp in files:
        try:
            content = fp.read_text(encoding="utf-8")
        except Exception:
            continue
        # 按 ^_^ 切分记录（备注字段值含 ^_^，下一记录紧跟）
        blocks = re.split(re.escape("^_^"), content)
        for block in blocks:
            block = block.strip()
            if not block:
                continue
            rec = _parse_school_block(block)
            if rec and rec.get("开设专业"):
                rec["_source_file"] = fp.name
                records.append(rec)
    return records


def _parse_school_block(block: str) -> Dict[str, Any]:
    """解析单条学校记录块"""
    rec: Dict[str, Any] = {}
    for line in block.splitlines():
        line = line.strip()
        if "：" in line:
            k, v = line.split("：", 1)
            key = k.strip()
            val = v.strip()
            if key == "备注" and val.startswith("^_^"):
                val = val[3:].strip()  # 去掉 ^_^ 前缀
            if key and key != "备注":
                rec[key] = val
    return rec


def load_questionnaire(path: Optional[Path] = None) -> List[Dict[str, Any]]:
    """加载问卷星 Excel 数据，返回行字典列表。优先用 pandas，失败则用 openpyxl"""
    path = path or QUESTIONNAIRE_XLSX
    if not path.exists():
        return []

    # 方式 1: pandas
    try:
        import pandas as pd

        df = pd.read_excel(path, sheet_name=0)
        df = df.fillna("")
        return df.to_dict(orient="records")
    except ImportError:
        pass
    except Exception:
        pass

    # 方式 2: openpyxl 回退
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            return []
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2):
            vals = []
            for c in row:
                v = c.value
                vals.append("" if v is None else str(v).strip())
            rows.append(dict(zip(headers, vals)))
        wb.close()
        return rows
    except ImportError:
        return []
    except Exception:
        return []


def load_questionnaire_columns(path: Optional[Path] = None) -> List[str]:
    """获取问卷星表头列名（用于调试或动态映射）"""
    path = path or QUESTIONNAIRE_XLSX
    if not path.exists():
        return []
    try:
        import pandas as pd

        df = pd.read_excel(path, sheet_name=0, nrows=1)
        return list(df.columns)
    except Exception:
        pass
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws:
            cols = [str(c.value).strip() if c.value else "" for c in ws[1]]
            wb.close()
            return cols
        wb.close()
    except Exception:
        pass
    return []
